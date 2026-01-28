from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Q, Count
from django.db import IntegrityError, transaction
from django.utils import timezone
from datetime import timedelta, datetime, date, time
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.forms.models import model_to_dict
from openpyxl.styles import Font, PatternFill, Alignment
import calendar
import openpyxl
import json
import uuid


# Imports locais
from .forms import ApontamentoForm
from .models import Apontamento, Projeto, Colaborador, Setor, Veiculo, CodigoCliente, CentroCusto, ApontamentoHistorico

# ==============================================================================
# 1. LÓGICA DE CONTROLE DE ACESSO (RBAC & HELPERS)
# ==============================================================================

def is_owner(user):
    return user.is_superuser

def check_group(user, group_name):
    return user.groups.filter(name=group_name).exists()

def is_coordenador(user):
    return check_group(user, 'COORDENADOR') or is_owner(user)

def is_administrativo(user):
    return check_group(user, 'ADMINISTRATIVO') or is_owner(user)

def is_gerente(user):
    return check_group(user, 'GESTOR') or is_owner(user)

def pode_fazer_rateio(user):
    """Regra: Coordenador, Administrativo e Owner podem ratear."""
    return is_coordenador(user) or is_administrativo(user) or is_owner(user)

def distribuir_horarios_com_gap(inicio, fim, qtd_obras):
    """Calcula horários sequenciais SEM INTERVALOS (Gap Zero)."""
    if qtd_obras <= 0: return []
    d = date(2000, 1, 1)
    dt_ini = datetime.combine(d, inicio)
    dt_fim = datetime.combine(d, fim)
    if dt_fim < dt_ini: dt_fim += timedelta(days=1)
    total_minutos = int((dt_fim - dt_ini).total_seconds() / 60)
    minutos_base = total_minutos // qtd_obras
    resto = total_minutos % qtd_obras
    intervalos = []
    tempo_atual = dt_ini
    for i in range(qtd_obras):
        duracao = minutos_base + (1 if i < resto else 0)
        if duracao < 1 and total_minutos > 0: duracao = 1 
        fim_obra = tempo_atual + timedelta(minutes=duracao)
        intervalos.append((tempo_atual.time(), fim_obra.time()))
        tempo_atual = fim_obra
    return intervalos

@login_required
def home_redirect_view(request):
    return redirect('produtividade:home_menu')

@login_required
def home_view(request):
    is_gestor = is_gerente(request.user)
    is_owner_user = is_owner(request.user)

    context = {
        'is_gestor': is_gestor,
        'is_owner': is_owner_user
    }
    return render(request, 'produtividade/home.html', context)

@login_required
def configuracoes_view(request):
    context = {
        'titulo': 'Configurações do Usuário', 
        'change_password_url': '/accounts/password_change/' }
    return render(request, 'produtividade/configuracoes.html', context)

# ==============================================================================
# 2. VIEWS DE OPERAÇÃO (CRIAR)
# ==============================================================================

@login_required
def apontamento_atividade_view(request):
    user_kwargs = {'user': request.user}
    
    if request.method == 'POST':
        form = ApontamentoForm(request.POST, **user_kwargs)
        if form.is_valid():
            apontamento = form.save(commit=False)
            apontamento.registrado_por = request.user
            apontamento.status_aprovacao = 'EM_ANALISE'
            apontamento.contagem_edicao = 0
            
            # --- Tratamento de Auxiliar/Veículo (Igual ao anterior) ---
            if form.cleaned_data.get('registrar_auxiliar'):
                apontamento.auxiliar = form.cleaned_data.get('auxiliar_selecao')
            else:
                apontamento.auxiliar = None

            if form.cleaned_data.get('registrar_veiculo'):
                selection = form.cleaned_data.get('veiculo_selecao')
                if selection == 'OUTRO':
                    apontamento.veiculo = None
                    apontamento.veiculo_manual_modelo = form.cleaned_data.get('veiculo_manual_modelo')
                    apontamento.veiculo_manual_placa = form.cleaned_data.get('veiculo_manual_placa')
                else:
                    try:
                        apontamento.veiculo = Veiculo.objects.get(pk=selection)
                        apontamento.veiculo_manual_modelo = None; apontamento.veiculo_manual_placa = None
                    except Veiculo.DoesNotExist: apontamento.veiculo = None
            else:
                apontamento.veiculo = None; apontamento.veiculo_manual_modelo = None; apontamento.veiculo_manual_placa = None

            # --- RATEIO ---
            extras_obras_str = form.cleaned_data.get('obras_extras_list')
            user_can_rateio = pode_fazer_rateio(request.user)
            
            is_rateio = user_can_rateio and (form.cleaned_data.get('registrar_multiplas_obras') or extras_obras_str)

            if is_rateio:
                agrupamento_uid = str(uuid.uuid4()) 

                principal_str = ""
                if apontamento.projeto: principal_str = f"P_{apontamento.projeto.id}"
                elif apontamento.codigo_cliente: principal_str = f"C_{apontamento.codigo_cliente.id}"
                
                lista_extras = [x.strip() for x in extras_obras_str.split(',') if x.strip()] if extras_obras_str else []
                todas_obras_raw = ([principal_str] + lista_extras) if principal_str else lista_extras

                if not todas_obras_raw:
                    apontamento.status_aprovacao = 'EM_ANALISE'
                    apontamento.save()
                    messages.success(request, "Registro salvo (único).")
                    return redirect('produtividade:novo_apontamento')

                horarios = distribuir_horarios_com_gap(apontamento.hora_inicio, apontamento.hora_termino, len(todas_obras_raw))
                aux_extras_str = form.cleaned_data.get('auxiliares_extras_list')
                ids_aux_list = [int(x) for x in aux_extras_str.split(',') if x.strip().isdigit()] if aux_extras_str else []
                contagem_sucesso = 0

                for idx, item_hibrido in enumerate(todas_obras_raw):
                    try:
                        if '_' not in item_hibrido: continue
                        prefixo, obj_id_str = item_hibrido.split('_')
                        obj_id = int(obj_id_str)

                        novo_registro = Apontamento()
                        novo_registro.colaborador = apontamento.colaborador
                        novo_registro.data_apontamento = apontamento.data_apontamento
                        novo_registro.local_execucao = apontamento.local_execucao
                        novo_registro.veiculo = apontamento.veiculo
                        novo_registro.veiculo_manual_modelo = apontamento.veiculo_manual_modelo
                        novo_registro.veiculo_manual_placa = apontamento.veiculo_manual_placa
                        novo_registro.auxiliar = apontamento.auxiliar
                        novo_registro.ocorrencias = apontamento.ocorrencias
                        novo_registro.em_plantao = apontamento.em_plantao
                        novo_registro.data_plantao = apontamento.data_plantao
                        novo_registro.dorme_fora = apontamento.dorme_fora
                        novo_registro.data_dorme_fora = apontamento.data_dorme_fora
                        novo_registro.latitude = apontamento.latitude
                        novo_registro.longitude = apontamento.longitude
                        novo_registro.registrado_por = request.user
                        novo_registro.status_aprovacao = 'EM_ANALISE'
                        novo_registro.contagem_edicao = 0
                        novo_registro.id_agrupamento = agrupamento_uid 

                        if idx < len(horarios):
                            novo_registro.hora_inicio = horarios[idx][0]
                            novo_registro.hora_termino = horarios[idx][1]
                        
                        if prefixo == 'P':
                            if not Projeto.objects.filter(pk=obj_id).exists(): continue
                            novo_registro.projeto_id = obj_id; novo_registro.codigo_cliente = None
                        elif prefixo == 'C':
                            if not CodigoCliente.objects.filter(pk=obj_id).exists(): continue
                            novo_registro.codigo_cliente_id = obj_id; novo_registro.projeto = None
                        
                        novo_registro.save()
                        contagem_sucesso += 1
                        if form.cleaned_data.get('registrar_auxiliar') and ids_aux_list:
                            novo_registro.auxiliares_extras.set(ids_aux_list)
                    except Exception as e:
                        print(f"Erro rateio: {e}")
                        continue
                
                if contagem_sucesso > 0: messages.success(request, f"Rateio realizado: {contagem_sucesso} registros.")
                else: messages.error(request, "Erro ao salvar rateio.")
                return redirect('produtividade:novo_apontamento')

            else:
                apontamento.status_aprovacao = 'EM_ANALISE'
                apontamento.save() 
                if form.cleaned_data.get('registrar_auxiliar'):
                    ids_string = form.cleaned_data.get('auxiliares_extras_list')
                    if ids_string:
                        ids_list = [int(x) for x in ids_string.split(',') if x.strip().isdigit()]
                        apontamento.auxiliares_extras.set(ids_list)
                    else:
                        apontamento.auxiliares_extras.clear()
                else:
                    apontamento.auxiliares_extras.clear()

                messages.success(request, f"Registro de {apontamento.colaborador} salvo com sucesso!")
                return redirect('produtividade:novo_apontamento')
    else:
        now_local = timezone.localtime(timezone.now())
        initial_data = {
            'data_apontamento': now_local.strftime('%d/%m/%Y'),
            'hora_inicio': now_local.strftime('%H:%M'),
        }
        form = ApontamentoForm(initial=initial_data, **user_kwargs)

    context = {
        'form': form,
        'titulo': 'Timesheet',
        'subtitulo': 'Preencha os dados de horário e local de trabalho.',
        'is_editing': False
    }
    return render(request, 'produtividade/apontamento_form.html', context)


# ==============================================================================
# 3. EDIÇÃO E HISTÓRICO DE APONTAMENTOS
# ==============================================================================

@login_required
def historico_apontamentos_view(request):
    """
    View de Listagem com filtros de data e permissões de visualização.
    """
    user = request.user
    
    # Eager Loading para evitar N+1 queries
    queryset = Apontamento.objects.select_related(
        'projeto', 'codigo_cliente', 'colaborador', 
        'veiculo', 'centro_custo', 'registrado_por'
    ).prefetch_related('auxiliares_extras').all()

    # --- Filtros de Data ---
    period = request.GET.get('period')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=6)
    current_period = '7'

    if period:
        try:
            days = int(period)
            start_date = end_date - timedelta(days=days - 1)
            current_period = period
            start_date_str = None
            end_date_str = None
        except ValueError:
            pass
    elif start_date_str and end_date_str:
        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            current_period = 'custom'
        except ValueError:
            pass

    queryset = queryset.filter(data_apontamento__gte=start_date, data_apontamento__lte=end_date)

    # --- Regras de Visualização ---
    if not is_owner(user) and not is_gerente(user):
        try:
            colab = Colaborador.objects.get(user_account=user)
            # Operacional vê o seu. Admin vê os do setor. (Implementar lógica Admin aqui se quiser refinar)
            queryset = queryset.filter(Q(registrado_por=user) | Q(colaborador=colab))
        except:
            queryset = queryset.filter(registrado_por=user)
        
        # Limita histórico para não-admins (segurança/performance)
        limit_date = timezone.now().date() - timedelta(days=30)
        if start_date < limit_date:
            start_date = limit_date
        queryset = queryset.filter(data_apontamento__gte=limit_date)

    # Processamento para exibição
    historico_lista = []

    total_segundos_geral = 0

    for item in queryset:
        # Formatação inteligente do Local
        if item.local_execucao == 'INT':
            local_tipo_display = "DENTRO DA OBRA"
            if item.projeto:
                p_cod = item.projeto.codigo if item.projeto.codigo else ""
                local_ref = f"{p_cod} - {item.projeto.nome}" if p_cod else f"{item.projeto.nome}"
            elif item.codigo_cliente:
                local_ref = f"{item.codigo_cliente.codigo} - {item.codigo_cliente.nome}"
            else:
                local_ref = "Obra/Cliente não informado"
        else:
            local_tipo_display = "FORA DO SETOR"
            local_ref = item.centro_custo.nome if item.centro_custo else "Atividade Externa"
            if item.projeto:
                local_ref += f" (Obra: {item.projeto.codigo})"
            elif item.codigo_cliente:
                local_ref += f" (CLIENTE: {item.codigo_cliente.codigo})"

        # Formatação inteligente do Veículo
        if item.veiculo: 
            veiculo_display = str(item.veiculo)
        elif item.veiculo_manual_placa: 
            veiculo_display = f"{item.veiculo_manual_modelo} - {item.veiculo_manual_placa} (Externo)"
        else: 
            veiculo_display = ""

        reg_user = item.registrado_por
        user_display = f"{reg_user.first_name} {reg_user.last_name}" if reg_user and reg_user.first_name else (reg_user.username if reg_user else "Sistema")

        # Horas totais do apontamento
        duracao_formatada = item.duracao_total_str

        base_dict = {
            'id': item.id,
            'data': item.data_apontamento,
            'local_ref': local_ref,
            'local_tipo': local_tipo_display,
            'inicio': item.hora_inicio,
            'termino': item.hora_termino,
            'duracao': duracao_formatada,
            'local_tipo': item.get_local_execucao_display(),
            'obs': item.ocorrencias,
            'registrado_em': item.data_registro,
            'registrado_por_str': user_display,
            'registrado_por_id': item.registrado_por.id if item.registrado_por else None,
            'em_plantao': item.em_plantao,
            'dorme_fora': item.dorme_fora,
            'motivo_ajuste': item.motivo_ajuste,
            'status_ajuste': item.status_ajuste,
            'status_aprovacao': item.status_aprovacao,
            'contagem_edicao': item.contagem_edicao,
            'pode_editar': (item.contagem_edicao < 1) or is_owner(user),
            'motivo_rejeicao': item.motivo_rejeicao,
            'latitude': item.latitude,
            'longitude': item.longitude,
        }

        # Adiciona linha principal (Colaborador)
        row_main = base_dict.copy()
        row_main.update({
            'nome': item.colaborador.nome_completo, 
            'cargo': item.colaborador.cargo, 
            'veiculo': veiculo_display, 
            'is_auxiliar': False
        })
        historico_lista.append(row_main)

        # Adiciona linhas de Auxiliares (se houver) para visualização expandida
        auxiliares_a_exibir = []
        if item.auxiliar: auxiliares_a_exibir.append(item.auxiliar)
        
        extras = item.auxiliares_extras.all()
        auxiliares_a_exibir.extend(extras)

        for aux in auxiliares_a_exibir:
            row_aux = base_dict.copy()
            row_aux.update({
                'nome': aux.nome_completo,
                'cargo': aux.cargo,
                'veiculo': "",
                'is_auxiliar': True
            })
            historico_lista.append(row_aux)

    total_horas_periodo = f"{total_segundos_geral // 3600:02d}:{(total_segundos_geral % 3600) // 60:02d}"

    context = {
        'titulo': "Histórico",
        'apontamentos_lista': historico_lista,
        'show_user_column': is_owner(user), 
        'is_owner': is_owner(user),
        'is_gestor': is_gerente(user),
        'current_period': current_period,
        'start_date_val': start_date.strftime('%Y-%m-%d'),
        'end_date_val': end_date.strftime('%Y-%m-%d'),
        'total_horas_periodo': total_horas_periodo,
    }
    return render(request, 'produtividade/historico_apontamentos.html', context)

@login_required
def editar_apontamento_view(request, pk):
    """
    Edição com controle de versão e limite de 1x.
    """
    apontamento = get_object_or_404(Apontamento, pk=pk)
    user = request.user

    if not is_owner(user) and apontamento.registrado_por != user:
        messages.error(request, "Acesso Negado: Você só pode editar seus próprios apontamentos.")
        return redirect('produtividade:historico_apontamentos')

    if apontamento.contagem_edicao >= 1 and not is_owner(user):
        messages.error(request, "Limite de edição atingido. Para correções, utilize a opção 'Solicitar Ajuste'.")
        return redirect('produtividade:historico_apontamentos')

    user_kwargs = {'user': request.user, 'instance': apontamento}

    if request.method == 'POST':
        dados_originais = model_to_dict(apontamento, exclude=['auxiliares_extras', 'user_account'])
        
        for k, v in dados_originais.items():
            if isinstance(v, (datetime, date, time)): 
                dados_originais[k] = v.isoformat()
            elif isinstance(v, timedelta): 
                dados_originais[k] = str(v)

        form = ApontamentoForm(request.POST, **user_kwargs)
        if form.is_valid():
            with transaction.atomic():
                ApontamentoHistorico.objects.create(
                    apontamento_original=apontamento,
                    dados_snapshot=dados_originais,
                    editado_por=user,
                    numero_edicao=apontamento.contagem_edicao + 1
                )

                obj = form.save(commit=False)
                obj.contagem_edicao += 1
                obj.status_aprovacao = 'EM_ANALISE'
                obj.motivo_rejeicao = None
                
                if not form.cleaned_data.get('registrar_auxiliar'): obj.auxiliar = None
                if not form.cleaned_data.get('registrar_veiculo'):
                    obj.veiculo = None; obj.veiculo_manual_modelo = None; obj.veiculo_manual_placa = None
                
                obj.save()

                if form.cleaned_data.get('registrar_auxiliar'):
                    ids_string = form.cleaned_data.get('auxiliares_extras_list')
                    if ids_string:
                        ids_list = [int(x) for x in ids_string.split(',') if x.strip().isdigit()]
                        obj.auxiliares_extras.set(ids_list)
                    else: obj.auxiliares_extras.clear()
                else: obj.auxiliares_extras.clear()

            messages.success(request, "Apontamento editado com sucesso! (Histórico salvo)")
            return redirect('produtividade:historico_apontamentos')
    else:
        initial_data = {}
        if apontamento.veiculo:
            initial_data['registrar_veiculo'] = True
            initial_data['veiculo_selecao'] = apontamento.veiculo.id
        elif apontamento.veiculo_manual_placa:
            initial_data['registrar_veiculo'] = True
            initial_data['veiculo_selecao'] = 'OUTRO'
        if apontamento.auxiliar:
            initial_data['registrar_auxiliar'] = True
            ids_list = list(apontamento.auxiliares_extras.values_list('id', flat=True))
            initial_data['auxiliares_extras_list'] = ",".join(map(str, ids_list))

        form = ApontamentoForm(initial=initial_data, **user_kwargs)

    context = {
        'form': form,
        'titulo': 'Editar Apontamento',
        'subtitulo': f'Editando registro (Versão {apontamento.contagem_edicao + 1})',
        'is_editing': True,
        'apontamento_id': pk
    }
    return render(request, 'produtividade/apontamento_form.html', context)

@login_required
def solicitar_ajuste_view(request, pk):
    """
    Permite que o usuário ou colaborador solicite um ajuste em um registro fechado.
    """
    apontamento = get_object_or_404(Apontamento, pk=pk)
    
    # Segurança: Só permite se o usuário for o dono do registro ou o colaborador vinculado
    is_autor = apontamento.registrado_por == request.user
    is_colaborador = False
    try:
        colab = Colaborador.objects.get(user_account=request.user)
        if apontamento.colaborador == colab:
            is_colaborador = True
    except Colaborador.DoesNotExist:
        pass

    if not (is_autor or is_colaborador or request.user.is_superuser):
         messages.error(request, "Você não tem permissão para solicitar ajuste neste registro.")
         return redirect('produtividade:historico_apontamentos')

    if request.method == 'POST':
        motivo = request.POST.get('motivo_texto')
        if motivo:
            apontamento.motivo_ajuste = motivo
            apontamento.status_aprovacao = 'SOLICITACAO_AJUSTE' 
            apontamento.status_ajuste = 'PENDENTE' 
            apontamento.save()
            messages.success(request, "Solicitação de ajuste enviada para a administração.")
        else:
            messages.warning(request, "É necessário descrever o motivo do ajuste.")
            
    return redirect('produtividade:historico_apontamentos')


@login_required
@user_passes_test(is_owner)
def excluir_apontamento_view(request, pk):
    """Exclusão de registro (Acesso Admin)."""
    apontamento = get_object_or_404(Apontamento, pk=pk)
    apontamento.delete()
    messages.success(request, "Apontamento excluído com sucesso.")
    return redirect('produtividade:historico_apontamentos')


@login_required
@user_passes_test(is_owner)
def aprovar_ajuste_view(request, pk):
    """Aprovação rápida de ajuste sem necessidade de edição."""
    apontamento = get_object_or_404(Apontamento, pk=pk)
    apontamento.status_ajuste = 'APROVADO'
    apontamento.save()
    messages.success(request, "Solicitação marcada como APROVADA.")
    return redirect('produtividade:historico_apontamentos')


# ==============================================================================
# 4. APIs AJAX & UTILITÁRIOS
# ==============================================================================

@login_required
def apontamento_sucesso_view(request):
    return render(request, 'produtividade/apontamento_sucesso.html')

@login_required
def get_projeto_info_ajax(request, projeto_id):
    projeto = get_object_or_404(Projeto, pk=projeto_id)
    return JsonResponse({'nome_projeto': projeto.nome})

@login_required
def get_colaborador_info_ajax(request, colaborador_id):
    colaborador = get_object_or_404(Colaborador, pk=colaborador_id)
    return JsonResponse({'cargo': colaborador.cargo})

@login_required
def get_auxiliares_ajax(request):
    auxs = Colaborador.objects.filter(cargo__in=['AUXILIAR TECNICO', 'OFICIAL DE SISTEMAS']).values('id', 'nome_completo')
    return JsonResponse({'auxiliares': list(auxs)})

@login_required
def get_centro_custo_info_ajax(request, cc_id):
    cc = get_object_or_404(CentroCusto, pk=cc_id)
    return JsonResponse({'permite_alocacao': cc.permite_alocacao})

@login_required
def get_calendar_status_ajax(request):
    """
    Retorna o status dos dias no calendário (preenchido, dorme_fora, etc.) para feedback visual.
    """
    try:
        month = int(request.GET.get('month'))
        year = int(request.GET.get('year'))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Parâmetros inválidos'}, status=400)
    
    user = request.user
    if is_owner(user): return JsonResponse({'is_owner': True, 'days': []})

    try:
        colaborador = Colaborador.objects.get(user_account=user)
    except Colaborador.DoesNotExist:
        return JsonResponse({'error': 'Colaborador não encontrado'}, status=400)

    _, num_days = calendar.monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)

    queryset = Apontamento.objects.filter(
        colaborador=colaborador, data_apontamento__gte=start_date, data_apontamento__lte=end_date
    ).values('data_apontamento', 'dorme_fora', 'em_plantao')
    
    # Mapeia dias com atividades
    dias_info = {}
    for entry in queryset:
        d_str = entry['data_apontamento'].strftime('%Y-%m-%d')
        if d_str not in dias_info:
            dias_info[d_str] = {
                'dorme_fora': entry['dorme_fora'],
                'em_plantao': entry['em_plantao']
            }
        else:
            if entry['dorme_fora']: dias_info[d_str]['dorme_fora'] = True
            if entry['em_plantao']: dias_info[d_str]['em_plantao'] = True
    
    days_data = []
    today = timezone.now().date()

    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        date_str = current_date.strftime('%Y-%m-%d')
        
        status = 'missing'
        has_dorme_fora = False
        has_em_plantao = False

        if date_str in dias_info:
            status = 'filled'
            has_dorme_fora = dias_info[date_str]['dorme_fora']
            has_em_plantao = dias_info[date_str]['em_plantao']
        elif current_date > today:
            status = 'future'
        
        days_data.append({
            'date': date_str, 
            'day': day, 
            'status': status,
            'has_dorme_fora': has_dorme_fora,
            'has_em_plantao': has_em_plantao
        })

    return JsonResponse({'is_owner': False, 'days': days_data})

@csrf_exempt
def api_dashboard_data(request):
    """
    API JSON para alimentar o Dashboard externo (PHP) ou interno.
    Agora protegida por API Key, igual à exportação.
    """
    # 1. Segurança via API Key
    api_key_esperada = getattr(settings, 'DJANGO_API_KEY', 'chave_secreta_123')
    token_recebido = request.headers.get('X-API-KEY')

    # Permite acesso se tiver a chave OU se for usuário logado no navegador
    if token_recebido != api_key_esperada and not request.user.is_authenticated:
         return JsonResponse({'erro': 'Acesso Negado'}, status=403)

    # 2. Definir o range de datas (Vamos pegar dados de HOJE)
    hoje = timezone.now().date()
    
    # 3. Buscar os apontamentos
    qs = Apontamento.objects.filter(data_apontamento=hoje).select_related('projeto', 'colaborador')

    # 4. Processar métricas
    total_registros = qs.count()
    total_segundos = 0
    projetos_ativos = {}
    colaboradores_ids = set()

    for a in qs:
        # Calcular horas (Termino - Inicio)
        if a.hora_inicio and a.hora_termino:
            dummy_date = date(2000, 1, 1)
            dt_inicio = datetime.combine(dummy_date, a.hora_inicio)
            dt_termino = datetime.combine(dummy_date, a.hora_termino)
            
            # Ajuste para virada de noite
            if dt_termino < dt_inicio:
                dt_termino += timedelta(days=1)
            
            diff = dt_termino - dt_inicio
            total_segundos += diff.total_seconds()

        # Contagem por Projeto
        nome_proj = "Outros"
        if a.local_execucao == 'INT':
             if a.projeto: nome_proj = a.projeto.nome
             elif a.codigo_cliente: nome_proj = f"Cliente {a.codigo_cliente.codigo}"
        else:
             if a.centro_custo: nome_proj = a.centro_custo.nome

        projetos_ativos[nome_proj] = projetos_ativos.get(nome_proj, 0) + 1
        
        # Colaboradores únicos
        if a.colaborador:
            colaboradores_ids.add(a.colaborador.nome_completo)

    # Converter segundos para Horas decimais
    total_horas = round(total_segundos / 3600, 2)

    # 5. Montar o JSON de resposta
    data = {
        'data_referencia': hoje.strftime('%d/%m/%Y'),
        'kpis': {
            'total_apontamentos': total_registros,
            'total_horas': total_horas,
            'colaboradores_ativos': len(colaboradores_ids),
        },
        'grafico_projetos': {
            'labels': list(projetos_ativos.keys()),
            'valores': list(projetos_ativos.values())
        },
        'lista_colaboradores': list(colaboradores_ids)
    }

    return JsonResponse(data)

# ==============================================================================
# 5. RELATÓRIOS (EXPORTAÇÃO EXCEL)
# ==============================================================================

@login_required
@user_passes_test(is_owner)
def exportar_relatorio_excel(request):
    """
    Gera um relatório consolidado em Excel para conferência de folha e custos.
    """
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    queryset = Apontamento.objects.select_related(
        'projeto', 'colaborador', 'veiculo', 'centro_custo', 'codigo_cliente'
    ).prefetch_related('auxiliares_extras').all().order_by('data_apontamento')
    
    if start_date_str and end_date_str:
        try:
            start = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            queryset = queryset.filter(data_apontamento__gte=start, data_apontamento__lte=end)
        except ValueError:
            pass

    # Setup do Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio de Horas"

    # Cabeçalho
    headers = [
        "Data", "Dia Semana", "Colaborador", "Cargo", "Tipo", 
        "Local (Obra/Setor)", "Código de Obra", "Código Cliente", 
        "Veículo", "Placa", "Hora Início", "Hora Fim", "Total Horas", 
        "Plantão", "Dorme Fora", "Observações", "Registrado Por", 'Latitude', 'Longitude'
    ]
    ws.append(headers)

    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def format_duration(inicio, fim):
        """Calcula a duração considerando virada de dia."""
        if not inicio or not fim: return "00:00:00"
        dummy_date = timezone.now().date()
        dt_inicio = timezone.datetime.combine(dummy_date, inicio)
        dt_fim = timezone.datetime.combine(dummy_date, fim)
        
        if dt_fim < dt_inicio:
            dt_fim += timedelta(days=1)

        diff = dt_fim - dt_inicio
        total_seconds = int(diff.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    dias_semana_pt = {
        0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira',
        3: 'Quinta-feira', 4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
    }

    for item in queryset:
        data_fmt = item.data_apontamento.strftime('%d/%m/%Y')
        dia_semana = dias_semana_pt[item.data_apontamento.weekday()]
        
        # Local
        local_nome = ""
        col_codigo_obra = ""
        col_codigo_cliente = ""

        if item.local_execucao == 'INT':
            tipo = "OBRA"
            if item.projeto:
                local_nome = item.projeto.nome
                col_codigo_obra = item.projeto.codigo
            elif item.codigo_cliente:
                local_nome = item.codigo_cliente.nome
                col_codigo_cliente = item.codigo_cliente.codigo
        else:
            tipo = "FORA DO SETOR"
            local_nome = item.centro_custo.nome if item.centro_custo else "Atividade Externa"
            if item.projeto: col_codigo_obra = item.projeto.codigo
            elif item.codigo_cliente: col_codigo_cliente = item.codigo_cliente.codigo

        if col_codigo_obra and len(col_codigo_obra) >= 5:
             col_codigo_cliente = col_codigo_obra[1:5]
        elif col_codigo_obra:
             col_codigo_cliente = col_codigo_obra

        # Veículo
        veiculo_nome_modelo = ""
        veiculo_placa_only = ""

        if item.veiculo:
            veiculo_nome_modelo = item.veiculo.descricao if item.veiculo.descricao else "Veículo da Frota"
            veiculo_placa_only = item.veiculo.placa
        elif item.veiculo_manual_modelo:
            veiculo_nome_modelo = item.veiculo_manual_modelo
            veiculo_placa_only = item.veiculo_manual_placa if item.veiculo_manual_placa else ""

        duracao_str = format_duration(item.hora_inicio, item.hora_termino)
        reg_por = item.registrado_por.username if item.registrado_por else "Sistema"

        # Apenas Status SIM/NÃO
        plantao_str = "SIM" if item.em_plantao else "NÃO"
        dorme_fora_str = "SIM" if item.dorme_fora else "NÃO"

        # Linha Principal (Colunas ajustadas)
        row_principal = [
            data_fmt, dia_semana, item.colaborador.nome_completo, item.colaborador.cargo,
            tipo, local_nome, col_codigo_obra, col_codigo_cliente, 
            veiculo_nome_modelo, veiculo_placa_only, item.hora_inicio, item.hora_termino, 
            duracao_str, plantao_str, dorme_fora_str, 
            item.ocorrencias, reg_por,
            item.latitude, item.longitude
        ]
        ws.append(row_principal)

        # Linhas Auxiliares (Carona)
        auxiliares = []
        if item.auxiliar:
            auxiliares.append(item.auxiliar)
        
        extras = list(item.auxiliares_extras.all())
        auxiliares.extend(extras)

        for aux in auxiliares:
            row_aux = [
                data_fmt, dia_semana, aux.nome_completo, aux.cargo,
                tipo, local_nome, col_codigo_obra, col_codigo_cliente, 
                "Carona", "", item.hora_inicio, item.hora_termino, 
                duracao_str, plantao_str, dorme_fora_str, 
                f"Auxiliar de: {item.colaborador.nome_completo}", reg_por
            ]
            ws.append(row_aux)

    # Auto-ajuste de largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Relatorio_Horas_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@csrf_exempt
def api_exportar_json(request):
    api_key_esperada = getattr(settings, 'DJANGO_API_KEY', 'chave_secreta_123')
    token_recebido = request.headers.get('X-API-KEY')
    if token_recebido != api_key_esperada: return JsonResponse({'erro': 'Acesso Negado'}, status=403)

    days = int(request.GET.get('days', 45))
    start_date = timezone.now().date() - timedelta(days=days)
    
    queryset = Apontamento.objects.select_related(
        'projeto', 'colaborador', 'veiculo', 'centro_custo', 'codigo_cliente'
    ).prefetch_related('auxiliares_extras').filter(
        data_apontamento__gte=start_date
    ).order_by('data_apontamento')

    dados_saida = []

    def fmt_hora(h): return h.strftime('%H:%M:%S') if h else None
    def fmt_data(d): return d.strftime('%Y-%m-%d') if d else None

    for item in queryset:
        local_nome = ""
        codigo_obra = None
        codigo_cliente = None
        
        if item.local_execucao == 'INT':
            tipo_str = "OBRA"
            if item.projeto:
                local_nome = item.projeto.nome
                codigo_obra = item.projeto.codigo
            elif item.codigo_cliente:
                local_nome = item.codigo_cliente.nome
                codigo_cliente = item.codigo_cliente.codigo
        else:
            tipo_str = "FORA DO SETOR"
            local_nome = item.centro_custo.nome if item.centro_custo else "Atividade Externa"
            if item.projeto: codigo_obra = item.projeto.codigo
            elif item.codigo_cliente: codigo_cliente = item.codigo_cliente.codigo

        if codigo_obra and len(str(codigo_obra)) >= 5:
             if not codigo_cliente: codigo_cliente = str(codigo_obra)[1:5]
        elif codigo_obra and not codigo_cliente:
             codigo_cliente = codigo_obra

        veiculo_nome = ""
        placa = ""
        if item.veiculo:
            veiculo_nome = item.veiculo.descricao
            placa = item.veiculo.placa
        elif item.veiculo_manual_modelo:
            veiculo_nome = item.veiculo_manual_modelo
            placa = item.veiculo_manual_placa

        base_obj = {
            'data': fmt_data(item.data_apontamento),
            'dia_semana': item.data_apontamento.weekday(), 
            'tipo': tipo_str,
            'local': local_nome,
            'codigo_obra': codigo_obra,
            'codigo_cliente': codigo_cliente,
            'hora_inicio': fmt_hora(item.hora_inicio),
            'hora_fim': fmt_hora(item.hora_termino), 
            'observacoes': item.ocorrencias,
            'registrado_por': item.registrado_por.username if item.registrado_por else 'Sistema',
            'dorme_fora': item.dorme_fora,
            'em_plantao': item.em_plantao,
            'status': item.status_ajuste or 'OK'
        }

        row_main = base_obj.copy()
        row_main.update({
            'colaborador': item.colaborador.nome_completo,
            'cargo': item.colaborador.cargo,
            'veiculo': veiculo_nome,
            'placa': placa,
            'is_auxiliar': False
        })
        dados_saida.append(row_main)

        auxiliares = []
        if item.auxiliar: auxiliares.append(item.auxiliar)
        auxiliares.extend(list(item.auxiliares_extras.all()))

        for aux in auxiliares:
            row_aux = base_obj.copy()
            row_aux.update({
                'colaborador': aux.nome_completo,
                'cargo': aux.cargo,
                'veiculo': 'Carona', 
                'placa': None,
                'is_auxiliar': True,
                # Auxiliares herdam o status do apontamento principal, 
                # mas geralmente não ganham plantão/dorme fora automaticamente (regra de negócio),
                # vamos manter false para evitar duplicidade de custo visual, ou true se a regra for essa.
                # Vou manter FALSE para caronas por segurança.
                'dorme_fora': True if item.dorme_fora else False, 
                'em_plantao': True if item.em_plantao else False, 
            })
            dados_saida.append(row_aux)

    return JsonResponse(dados_saida, safe=False)

# ==============================================================================
# 6. APROVAÇÃO DE AJUSTES (GERENTE)
# ==============================================================================

@login_required
@user_passes_test(is_gerente)
def aprovacao_dashboard_view(request):
    """
    Lista de pendências para o Gerente.
    Se for Owner (Superuser), vê tudo.
    Se for Gestor Comum, vê apenas colaboradores dos seus setores.
    """
    if is_owner(request.user):
        pendentes = Apontamento.objects.filter(
            status_aprovacao='EM_ANALISE'
        ).select_related('colaborador', 'projeto', 'centro_custo').order_by('data_apontamento')
        
    else:
        try:
            gerente = Colaborador.objects.get(user_account=request.user)
            meus_setores = gerente.setores_gerenciados.all()
            
            pendentes = Apontamento.objects.filter(
                status_aprovacao='EM_ANALISE',
                colaborador__setor__in=meus_setores
            ).exclude(colaborador=gerente).select_related('colaborador', 'projeto', 'centro_custo').order_by('data_apontamento')
            
        except Colaborador.DoesNotExist:
            messages.error(request, "Seu usuário não está vinculado a um cadastro de Colaborador/Gestor.")
            return redirect('produtividade:home_menu')

    context = {
        'pendentes': pendentes,
        'titulo': 'Central de Aprovações'
    }
    return render(request, 'produtividade/aprovacao_dashboard.html', context)


@login_required
@user_passes_test(is_gerente)
def analise_apontamento_view(request, pk):
    """
    Tela detalhada para comparar a versão anterior com a atual (Diff Completo).
    """
    apontamento = get_object_or_404(Apontamento, pk=pk)
    
    # --- HELPER DEFINIDO NO INÍCIO (CORREÇÃO) ---
    def item_time_str(t): 
        return t.strftime('%H:%M') if t else ""

    # Pega a ÚLTIMA versão do histórico (a imediatamente anterior à atual)
    historico = ApontamentoHistorico.objects.filter(apontamento_original=apontamento).order_by('-numero_edicao').first()
    
    diff_data = []
    tem_alteracao = False

    if historico:
        dados_antigos = historico.dados_snapshot
        
        # --- FUNÇÕES AUXILIARES DE FORMATAÇÃO ---
        def format_bool(val):
            return "SIM" if val else "NÃO"

        def format_none(val):
            return val if val else "-"

        def get_fk_name(ModelClass, pk):
            if not pk: return "-"
            try:
                return str(ModelClass.objects.get(pk=pk))
            except:
                return f"(ID: {pk} removido)"

        # --- 1. COMPARAÇÃO DE HORÁRIOS ---
        # Formata strings de hora para comparar (snapshot vem como string ISO ou HH:MM:SS)
        h_ini_old = str(dados_antigos.get('hora_inicio', ''))[:5]
        h_ini_new = item_time_str(apontamento.hora_inicio) # Agora funciona!
        if h_ini_old != h_ini_new:
            diff_data.append({'campo': 'Hora Início', 'antes': h_ini_old, 'depois': h_ini_new, 'icon': 'clock'})

        h_fim_old = str(dados_antigos.get('hora_termino', ''))[:5]
        h_fim_new = item_time_str(apontamento.hora_termino)
        if h_fim_old != h_fim_new:
            diff_data.append({'campo': 'Hora Término', 'antes': h_fim_old, 'depois': h_fim_new, 'icon': 'clock'})

        # --- 2. COMPARAÇÃO DE LOCAL/PROJETO ---
        local_old = dados_antigos.get('local_execucao')
        local_new = apontamento.local_execucao
        if local_old != local_new:
            mapa = {'INT': 'DENTRO DA OBRA', 'EXT': 'FORA DA OBRA'}
            diff_data.append({'campo': 'Local', 'antes': mapa.get(local_old, local_old), 'depois': mapa.get(local_new, local_new), 'icon': 'map'})

        # Comparação de Projeto (ID vs ID)
        proj_old_id = dados_antigos.get('projeto')
        proj_new_id = apontamento.projeto.id if apontamento.projeto else None
        if proj_old_id != proj_new_id:
            nome_old = get_fk_name(Projeto, proj_old_id)
            nome_new = str(apontamento.projeto) if apontamento.projeto else "-"
            diff_data.append({'campo': 'Projeto/Obra', 'antes': nome_old, 'depois': nome_new, 'icon': 'briefcase'})

        # Comparação de Cliente (ID vs ID)
        cli_old_id = dados_antigos.get('codigo_cliente')
        cli_new_id = apontamento.codigo_cliente.id if apontamento.codigo_cliente else None
        if cli_old_id != cli_new_id:
            nome_old = get_fk_name(CodigoCliente, cli_old_id)
            nome_new = str(apontamento.codigo_cliente) if apontamento.codigo_cliente else "-"
            diff_data.append({'campo': 'Cliente', 'antes': nome_old, 'depois': nome_new, 'icon': 'user'})

        # --- 3. COMPARAÇÃO DE VEÍCULO (Complexo: ID vs Manual) ---
        # Veículo Cadastrado
        veic_old_id = dados_antigos.get('veiculo')
        veic_new_id = apontamento.veiculo.id if apontamento.veiculo else None
        
        # Veículo Manual
        veic_man_placa_old = dados_antigos.get('veiculo_manual_placa')
        veic_man_placa_new = apontamento.veiculo_manual_placa

        # Lógica: Se mudou ID ou mudou Texto Manual
        if veic_old_id != veic_new_id:
            nome_old = get_fk_name(Veiculo, veic_old_id)
            nome_new = str(apontamento.veiculo) if apontamento.veiculo else "-"
            diff_data.append({'campo': 'Veículo (Frota)', 'antes': nome_old, 'depois': nome_new, 'icon': 'truck'})
        
        if str(veic_man_placa_old) != str(veic_man_placa_new):
             diff_data.append({'campo': 'Veículo (Externo/Placa)', 'antes': format_none(veic_man_placa_old), 'depois': format_none(veic_man_placa_new), 'icon': 'truck'})

        # --- 4. COMPARAÇÃO DE ADICIONAIS (Booleans) ---
        if dados_antigos.get('em_plantao') != apontamento.em_plantao:
            diff_data.append({
                'campo': 'Em Plantão?', 
                'antes': format_bool(dados_antigos.get('em_plantao')), 
                'depois': format_bool(apontamento.em_plantao),
                'icon': 'siren'
            })

        if dados_antigos.get('dorme_fora') != apontamento.dorme_fora:
            diff_data.append({
                'campo': 'Dorme Fora?', 
                'antes': format_bool(dados_antigos.get('dorme_fora')), 
                'depois': format_bool(apontamento.dorme_fora),
                'icon': 'moon'
            })

        # --- 5. COMPARAÇÃO DE TEXTOS ---
        obs_old = str(dados_antigos.get('ocorrencias', '') or '').strip()
        obs_new = str(apontamento.ocorrencias or '').strip()
        if obs_old != obs_new:
            diff_data.append({'campo': 'Observações', 'antes': obs_old, 'depois': obs_new, 'icon': 'pencil'})

        # --- 6. COMPARAÇÃO DE CENTRO DE CUSTO ---
        cc_old_id = dados_antigos.get('centro_custo')
        cc_new_id = apontamento.centro_custo.id if apontamento.centro_custo else None
        if cc_old_id != cc_new_id:
            nome_old = get_fk_name(CentroCusto, cc_old_id)
            nome_new = str(apontamento.centro_custo) if apontamento.centro_custo else "-"
            diff_data.append({'campo': 'Centro de Custo', 'antes': nome_old, 'depois': nome_new, 'icon': 'map'})

        # --- 7. COMPARAÇÃO DE MODELO VEÍCULO (MANUAL) ---
        mod_old = str(dados_antigos.get('veiculo_manual_modelo') or '')
        mod_new = str(apontamento.veiculo_manual_modelo or '')
        if mod_old != mod_new:
            diff_data.append({'campo': 'Modelo Veículo (Manual)', 'antes': format_none(mod_old), 'depois': format_none(mod_new), 'icon': 'truck'})

        # --- 8. COMPARAÇÃO DE AUXILIAR PRINCIPAL ---
        aux_old_id = dados_antigos.get('auxiliar')
        aux_new_id = apontamento.auxiliar.id if apontamento.auxiliar else None
        if aux_old_id != aux_new_id:
            nome_old = get_fk_name(Colaborador, aux_old_id)
            nome_new = str(apontamento.auxiliar.nome_completo) if apontamento.auxiliar else "-"
            diff_data.append({'campo': 'Auxiliar Principal', 'antes': nome_old, 'depois': nome_new, 'icon': 'user'})

        # --- 9. COMPARAÇÃO DE DATA DO REGISTRO ---
        data_old_str = str(dados_antigos.get('data_apontamento'))
        data_new_str = apontamento.data_apontamento.strftime('%Y-%m-%d')
        if data_old_str != data_new_str:
             d_old_fmt = datetime.strptime(data_old_str, '%Y-%m-%d').strftime('%d/%m/%Y')
             d_new_fmt = apontamento.data_apontamento.strftime('%d/%m/%Y')
             diff_data.append({'campo': 'Data do Registro', 'antes': d_old_fmt, 'depois': d_new_fmt, 'icon': 'calendar'})

        if diff_data: tem_alteracao = True   

    context = {
        'apontamento': apontamento,
        'historico': historico,
        'diffs': diff_data,
        'tem_alteracao': tem_alteracao,
        'duracao_total': apontamento.duracao_total_str if hasattr(apontamento, 'duracao_total_str') else "Calculando...",
        'usuario_editor': historico.editado_por if historico else None
    }
    return render(request, 'produtividade/aprovacao_analise.html', context)


@login_required
@user_passes_test(is_gerente)
def processar_aprovacao_view(request, pk):
    if request.method != 'POST': 
        return redirect('produtividade:aprovacao_dashboard')
    
    apontamento = get_object_or_404(Apontamento, pk=pk)
    acao = request.POST.get('acao')
    motivo = request.POST.get('motivo_rejeicao', '').strip()

    if not motivo:
        messages.error(request, "É obrigatório inserir um comentário/motivo para finalizar a análise.")
        return redirect('produtividade:analise_apontamento', pk=pk)

    if acao == 'APROVAR':
        apontamento.status_aprovacao = 'APROVADO'
        apontamento.motivo_rejeicao = motivo 
        messages.success(request, f"Registro APROVADO com sucesso.")
    
    elif acao == 'REJEITAR':
        apontamento.status_aprovacao = 'REJEITADO'
        apontamento.motivo_rejeicao = motivo
        messages.warning(request, f"Registro REJEITADO. O colaborador foi notificado.")

    apontamento.save()
    
    return redirect('produtividade:aprovacao_dashboard')